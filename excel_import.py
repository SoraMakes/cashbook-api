#!/usr/bin/python3

import openpyxl
import requests
import datetime
import logging
import sys

# --- Configuration ---
config = {
    'api_endpoint': 'http://localhost:8097',  # Replace with your API endpoint
    'username': 'test',          # Replace with your username
    'password': 'test',          # Replace with your password
    'excel_file': '20241029-18937-umsatz.xlsx',  # Replace with your Excel file path
    'column_mapping': {
        'category': 'category',  # Excel header: API field
        'Betrag': 'amount',
        'Buchungstag': 'date',
        'Verwendungszweck': 'description',
        'Beguenstigter/Zahlungspflichtiger': 'recipient_sender',
        'is_income': 'is_income',
        'no_invoice': 'no_invoice',
    }
}


logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('app.log', mode='a')
    ]
)

# Set the level for the StreamHandler to INFO
logging.getLogger().handlers[0].setLevel(logging.INFO)



def login(username, password):
    url = f"{config['api_endpoint']}/api/login"
    data = {'username': username, 'password': password}
    response = requests.post(url, json=data)
    response.raise_for_status()  # Raise an exception for bad status codes
    return response.json()  # Assuming the token is in the response JSON

def load_categories(token):
    url = f"{config['api_endpoint']}/api/categories"
    headers = {'Authorization': f'Bearer {token}'}
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    categories = {category['name']: category['id'] for category in response.json()}
    return categories

def load_excel_data(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(values_only=True))  # Get the header row
    column_indices = {}
    for i, cell_value in enumerate(header_row):
        if cell_value in config['column_mapping']:
            column_indices[config['column_mapping'][cell_value]] = i
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    # log the column indices
    logging.info(f"Column indices: {column_indices}")

    return data, column_indices, header_row

# --- Field mappings and transformations ---
def transform_data(row, categories, column_indices):
    try:
        date = row[column_indices['date']].strftime('%Y-%m-%dT%H:%M:%S')
        description = row[column_indices['description']]
        recipient_sender = row[column_indices['recipient_sender']]
        amount = int(abs(float(row[column_indices['amount']])) * 100)
        is_income = True if row[column_indices['is_income']].lower() in ('ja', 'yes', 'true') else False
        no_invoice = True if row[column_indices['no_invoice']].lower() in ('ja', 'yes', 'true') else False
        category_id = categories.get(row[column_indices['category']])

        # Validation - check if all required fields are present
        if not category_id:
            logging.warning(f"Category \"{row[column_indices['category']]}\" not found")
            return None

        missing_fields = []
        if not date:
            missing_fields.append('date')
        if not description:
            missing_fields.append('description')
        if not recipient_sender:
            missing_fields.append('recipient_sender')
        if not amount:
            missing_fields.append('amount')

        if missing_fields:
            raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

        entry = {
            'category_id': category_id,
            'amount': amount,
            'is_income': is_income,
            'recipient_sender': recipient_sender,
            'payment_method': 'bank_transfer',
            'description': description,
            'no_invoice': no_invoice,
            'date': date,
        }
        return entry
    except Exception as e:
        if not all(cell is None for cell in row):  # do not log error for empty rows
            logging.error(f"Error transforming data for row {row}: {e}")
        return None

def send_entry(entry, token):
    url = f"{config['api_endpoint']}/api/entries"
    headers = {'Authorization': f'Bearer {token}'}  # Include the token in the header
    response = requests.post(url, json=entry, headers=headers)
    response.raise_for_status()
    return response.json()

def save_entries_to_excel(filename, entries, header_row):
    logging.debug(f"Saving {len(entries)} entries to {filename}")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(header_row)  # Add the header row
    for entry in entries:
        sheet.append(entry)
    logging.info(f"Saving to {filename}")
    workbook.save(filename)

def process_entries():
    try:
        token = login(config['username'], config['password'])['token']  # Get the token
        categories = load_categories(token)
        excel_data, column_indices, header_row = load_excel_data(config['excel_file'])

        successful_entries = []
        failed_entries = []

        for row in excel_data:
            entry = transform_data(row, categories, column_indices)
            if entry:
                try:
                    response = send_entry(entry, token)
                    if 'amount' in response and isinstance(response['amount'], int):
                        successful_entries.append(row)
                        # Log successful entry
                        recipient_sender_log = entry['recipient_sender'][:40] + '...' if len(entry['recipient_sender']) > 40 else entry['recipient_sender']
                        description_log = entry['description'][:60] + '...' if len(entry['description']) > 60 else entry['description']
                        logging.debug(f"Entry processed successfully: recipient_sender={recipient_sender_log}, description={description_log}")
                    else:
                        failed_entries.append(row)
                        logging.error(f"Invalid API response: {response}")
                except Exception as e:
                    failed_entries.append(row)
                    logging.error(f"Error sending entry: {e}")
            else:
                failed_entries.append(row)  # Append original row data if transformation fails

        # --- Save successful and failed entries to Excel files ---
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        save_entries_to_excel(f"{timestamp}_successful.xlsx", successful_entries, header_row)
        save_entries_to_excel(f"{timestamp}_failed.xlsx", failed_entries, header_row)

    except Exception as e:
        logging.error(f"An error occurred: {e}")

if __name__ == "__main__":
    process_entries()
